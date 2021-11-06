import yfinance as yf
from pandas import Series
from constants import stocks, indexSymbolForYfinance
from math import log,exp
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from os import path ,remove


def getReturnsForData(closeSeries):

    skip = 1
    prevdata = None

    dailyReturnlist = []
    for data in closeSeries:

        if skip == 1:
            skip = 0
            prevdata = data
            continue

        dailyReturnlist.append((log(data/prevdata))*100)
        prevdata = data

    return dailyReturnlist

def addHeaderToExcel(workSheet,expiry):
    workSheet.append(['S.No',
                      'Stock Name',
                      'Current Price',
                      'Daily Volatility',
                      'Volatility after {} days'.format(expiry),
                      '1 SD (LL)',
                      '1 SD (HH)',
                      '2 SD (LL)',
                      '2 SD (HH)',
                      ])
    cellRange = workSheet['A1':'I1']
    redFill = PatternFill(start_color='FFFFFF00',
                          end_color='FFFFFF00',
                          fill_type='solid')

    for cell in cellRange[0]:
        cell.fill = redFill


def computeSTDandAddtoExcel(stock, data ,No_of_days_to_Expire, workSheet, sNo):

    # print(data.iloc[-1], data.__class__)
    # print(data[-9:])

    dailyReturns = Series(getReturnsForData(data))

    # print(dailyReturns,len(dailyReturns ) ,end='\n')
    mean = dailyReturns.mean()
    standardDeviation = dailyReturns.std()
    # print(mean)
    # print(standardDeviation)

    mean_at_expiry = mean * No_of_days_to_Expire
    std_at_expiry = standardDeviation * (pow(No_of_days_to_Expire,0.5))
    # print(mean_at_expiry)
    # print(std_at_expiry)

    lowerLimit_1SD = (exp((mean_at_expiry - std_at_expiry)/100)) * data.iloc[-1]
    upperLimit_1SD = (exp((mean_at_expiry + std_at_expiry)/100)) * data.iloc[-1]

    lowerLimit_2SD = (exp((mean_at_expiry - (2 * std_at_expiry))/100)) * data.iloc[-1]
    upperLimit_2SD = (exp((mean_at_expiry + (2 * std_at_expiry))/100)) * data.iloc[-1]

    workSheet.append([sNo,
                      stock,
                      data.iloc[-1],
                      standardDeviation,
                      std_at_expiry,
                      lowerLimit_1SD,
                      upperLimit_1SD,
                      lowerLimit_2SD,
                      upperLimit_2SD])

    # fp.write("******************************************************\n")
    # fp.write("Stock Name       : {}\n".format(stock))
    # fp.write("Current Price    : {}\n".format(data.iloc[-1]))
    # fp.write("Daily Volatility : {}\n".format(standardDeviation))
    # fp.write("Volatily after {} days: {}\n".format(No_of_days_to_Expire,std_at_expiry))
    # fp.write("Range for 1SD    : {} - {}\n".format(lowerLimit_1SD,upperLimit_1SD))
    # fp.write("Range for 2SD    : {} - {}\n".format(lowerLimit_2SD,upperLimit_2SD))
    # fp.write("******************************************************\n")

if __name__ == "__main__":

    No_of_days_to_Expire = int(input("Number of days to expire : "))

    fileName = "./standardDeviation_data({} days).xlsx".format(No_of_days_to_Expire)

    if (path.exists(fileName) == True):
        remove(fileName)

    wb = Workbook()
    ws = wb.active

    addHeaderToExcel(ws,No_of_days_to_Expire)
    count = 1

    for index in indexSymbolForYfinance:
        data = yf.download(indexSymbolForYfinance[index], period="3y")["Close"]
        computeSTDandAddtoExcel(index, data,No_of_days_to_Expire, ws, count)
        count+=1

    for stock in stocks:
        data = yf.download(stocks[stock]+".NS" , period="3y")["Close"]
        computeSTDandAddtoExcel(stock,data,No_of_days_to_Expire, ws,count)
        count += 1

    wb.save(fileName)

    wb.close()



            
            




