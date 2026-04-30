import yfinance as yf
from pandas import Series
from common.constants import stocks, indexSymbolForYfinance
from math import log,exp
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from os import path ,remove


def getReturnsForData(closeSeries):

    skip = 1
    prevdata = None

    dailyReturnlist = []
    for data in closeSeries:

        if skip == 1:
            skip = 0
            prevdata = data
            continue

        dailyReturnlist.append((log(data/prevdata))*100)
        prevdata = data

    return dailyReturnlist

def addHeaderToExcel(workSheet,expiry):
    workSheet.append(['S.No',
                      'Stock Name',
                      'Current Price',
                      'Average Returns',
                      'Average Returns for {} days'.format(expiry),
                      'Daily Volatility',
                      'Volatility after {} days'.format(expiry),
                      '1 SD (LL)',
                      '1 SD (HH)',
                      '2 SD (LL)',
                      '2 SD (HH)',
                      ])
    cellRange = workSheet['A1':'K1']
    redFill = PatternFill(start_color='FFFFFF00',
                          end_color='FFFFFF00',
                          fill_type='solid')

    for cell in cellRange[0]:
        cell.fill = redFill


def computeSTDandAddtoExcel(stock, data ,No_of_days_to_Expire, workSheet, sNo):

    dailyReturns = Series(getReturnsForData(data))

    mean = dailyReturns.mean()
    standardDeviation = dailyReturns.std()

    mean_at_expiry = mean * No_of_days_to_Expire
    std_at_expiry = standardDeviation * (pow(No_of_days_to_Expire,0.5))

    lowerLimit_1SD = (exp((mean_at_expiry - std_at_expiry)/100)) * data.iloc[-1]
    upperLimit_1SD = (exp((mean_at_expiry + std_at_expiry)/100)) * data.iloc[-1]

    lowerLimit_2SD = (exp((mean_at_expiry - (2 * std_at_expiry))/100)) * data.iloc[-1]
    upperLimit_2SD = (exp((mean_at_expiry + (2 * std_at_expiry))/100)) * data.iloc[-1]

    workSheet.append([sNo,
                      stock,
                      data.iloc[-1],
                      mean,
                      mean_at_expiry,
                      standardDeviation,
                      std_at_expiry,
                      lowerLimit_1SD,
                      upperLimit_1SD,
                      lowerLimit_2SD,
                      upperLimit_2SD])

if __name__ == "__main__":

    No_of_days_to_Expire = int(input("Number of days to expire : "))

    fileName = "./standardDeviation_data({} days).xlsx".format(No_of_days_to_Expire)

    if (path.exists(fileName) == True):
        remove(fileName)

    wb = Workbook()
    ws = wb.active

    addHeaderToExcel(ws,No_of_days_to_Expire)
    count = 1

    for index in indexSymbolForYfinance:
        data = yf.download(indexSymbolForYfinance[index], period="3y")["Close"]
        computeSTDandAddtoExcel(index, data,No_of_days_to_Expire, ws, count)
        count+=1

    for stock in stocks:
        data = yf.download(stocks[stock]+".NS" , period="3y")["Close"]
        computeSTDandAddtoExcel(stock,data,No_of_days_to_Expire, ws,count)
        count += 1

    wb.save(fileName)

    wb.close()



            
            




